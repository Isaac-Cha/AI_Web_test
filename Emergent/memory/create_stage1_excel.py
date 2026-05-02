from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def main() -> None:
  out_path = Path(r"d:\Aiproject1\EA\website\Emergent\memory\第一阶段-内容录入模板.xlsx")
  out_path.parent.mkdir(parents=True, exist_ok=True)

  headers = ["编号", "排序", "板块", "标题", "展现形式", "描述", "明细"]
  sheet_names = ["合作平台", "EA", "指标", "教学", "加入我们"]

  wb = Workbook()
  wb.remove(wb.active)

  header_font = Font(bold=True, color="FFFFFF")
  header_fill = PatternFill("solid", fgColor="111827")
  header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
  content_alignment = Alignment(vertical="top", wrap_text=True)
  widths = [10, 8, 18, 26, 18, 34, 42]

  for title in sheet_names:
    ws = wb.create_sheet(title=title)
    ws.append(headers)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

    for col_idx in range(1, len(headers) + 1):
      cell = ws.cell(row=1, column=col_idx)
      cell.font = header_font
      cell.fill = header_fill
      cell.alignment = header_alignment

    for i, w in enumerate(widths, start=1):
      ws.column_dimensions[get_column_letter(i)].width = w

    for col_idx in range(1, len(headers) + 1):
      for row_idx in range(2, 202):
        ws.cell(row=row_idx, column=col_idx).alignment = content_alignment

  wb.save(out_path)
  print(out_path)


if __name__ == "__main__":
  main()

